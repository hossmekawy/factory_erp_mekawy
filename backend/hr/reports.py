"""Weekly attendance report: Sat→Thu work week, Friday off.

Built from raw AttendanceLog punches: per employee per day the first punch is
check-in, the last is check-out; hours = the span between them.
"""
import datetime
from collections import defaultdict
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from devices.models import AttendanceLog
from .models import Employee, WorkSchedule

DAY_NAMES_AR = {
    5: "السبت", 6: "الأحد", 0: "الاثنين", 1: "الثلاثاء",
    2: "الأربعاء", 3: "الخميس", 4: "الجمعة",
}


def week_start_for(day: datetime.date) -> datetime.date:
    """Return the Saturday on or before `day` (weeks run Sat→Fri)."""
    offset = (day.weekday() - 5) % 7
    return day - datetime.timedelta(days=offset)


def build_weekly_report(anchor: datetime.date) -> dict:
    schedule = WorkSchedule.get_solo()
    start = week_start_for(anchor)
    all_days = [start + datetime.timedelta(days=i) for i in range(7)]
    workdays = [d for d in all_days if d.weekday() not in schedule.weekend_days]
    today = timezone.localdate()

    tz = timezone.get_current_timezone()
    range_start = datetime.datetime.combine(all_days[0], datetime.time.min, tzinfo=tz)
    range_end = datetime.datetime.combine(
        all_days[-1] + datetime.timedelta(days=1), datetime.time.min, tzinfo=tz
    )

    punches = defaultdict(list)  # (employee_code, date) -> [datetime,...]
    logs = AttendanceLog.objects.filter(
        timestamp__gte=range_start, timestamp__lt=range_end
    ).values_list("employee_code", "timestamp")
    for code, ts in logs:
        local = timezone.localtime(ts)
        punches[(code, local.date())].append(local)

    employees = []
    for emp in Employee.objects.filter(is_active=True).select_related("department"):
        days = {}
        present = absent = 0
        total_hours = 0.0
        for day in workdays:
            # NOTE: this first-punch/last-punch rule is duplicated in
            # hr/attendance.py, which the cutting module calls per employee.
            # The duplication is deliberate: this report is in daily use and
            # correct, and rewriting it on top of the new helper buys nothing
            # but risk. Change both together if the rule itself ever changes.
            cell = {"date": day.isoformat(), "status": "", "in": None, "out": None, "hours": 0}
            times = sorted(punches.get((emp.employee_code, day), []))
            if times:
                first, last = times[0], times[-1]
                hours = round((last - first).total_seconds() / 3600, 2) if last > first else 0
                cell["in"] = first.strftime("%H:%M")
                cell["out"] = last.strftime("%H:%M") if last > first else None
                cell["hours"] = hours
                cell["status"] = "present"
                present += 1
                total_hours += hours
            elif emp.hire_date and day < emp.hire_date:
                cell["status"] = "not_hired"
            elif day > today:
                cell["status"] = "future"
            else:
                cell["status"] = "absent"
                absent += 1
            days[day.isoformat()] = cell
        employees.append(
            {
                "id": emp.id,
                "employee_code": emp.employee_code,
                "full_name": emp.full_name,
                "department": emp.department.name if emp.department else "",
                "days": days,
                "totals": {
                    "present": present,
                    "absent": absent,
                    "hours": round(total_hours, 2),
                },
            }
        )

    return {
        "week_start": all_days[0].isoformat(),
        "week_end": all_days[-1].isoformat(),
        "workdays": [
            {"date": d.isoformat(), "name": DAY_NAMES_AR[d.weekday()]} for d in workdays
        ],
        "schedule": {
            "work_start": schedule.work_start.strftime("%H:%M"),
            "work_end": schedule.work_end.strftime("%H:%M"),
        },
        "employees": employees,
    }


def weekly_report_xlsx(report: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير الأسبوع"
    ws.sheet_view.rightToLeft = True

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF", name="Arial")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:K1")
    title = ws.cell(row=1, column=1)
    title.value = (
        f"MR.Mekawy Factory ERP — تقرير الحضور الأسبوعي "
        f"من {report['week_start']} إلى {report['week_end']}"
    )
    title.font = Font(bold=True, size=14, name="Arial")
    title.alignment = center

    headers = ["الكود", "الاسم", "القسم"]
    for d in report["workdays"]:
        headers.append(f"{d['name']}\n{d['date'][5:]}")
    headers += ["أيام الحضور", "أيام الغياب", "إجمالي الساعات"]

    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=header)
        c.fill = head_fill
        c.font = head_font
        c.alignment = center
        c.border = border

    absent_fill = PatternFill("solid", fgColor="FDE9E9")
    present_fill = PatternFill("solid", fgColor="EAF6EA")

    row = 4
    for emp in report["employees"]:
        ws.cell(row=row, column=1, value=emp["employee_code"]).alignment = center
        ws.cell(row=row, column=2, value=emp["full_name"])
        ws.cell(row=row, column=3, value=emp["department"]).alignment = center
        col = 4
        for d in report["workdays"]:
            cell = emp["days"][d["date"]]
            c = ws.cell(row=row, column=col)
            c.alignment = center
            if cell["status"] == "present":
                if cell["out"]:
                    c.value = f"{cell['in']} - {cell['out']}\n({cell['hours']} س)"
                else:
                    c.value = f"{cell['in']}\n(بصمة واحدة)"
                c.fill = present_fill
            elif cell["status"] == "absent":
                c.value = "غياب"
                c.fill = absent_fill
            else:
                c.value = "—"
            col += 1
        ws.cell(row=row, column=col, value=emp["totals"]["present"]).alignment = center
        ws.cell(row=row, column=col + 1, value=emp["totals"]["absent"]).alignment = center
        ws.cell(row=row, column=col + 2, value=emp["totals"]["hours"]).alignment = center
        for c_idx in range(1, col + 3):
            ws.cell(row=row, column=c_idx).border = border
        row += 1

    widths = [10, 28, 14] + [16] * len(report["workdays"]) + [12, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(4, row):
        ws.row_dimensions[r].height = 30

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
