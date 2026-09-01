"""Cutting reports (SRS section 9), in the same shape as hr/reports.py:
one builder returns a plain dict, and the same dict feeds JSON, Excel and PDF.

Every report excludes backfilled lays by default. A backfilled lay is a
notebook page from before the system existed; mixing it into an operational
report distorts the very numbers the report exists to watch (SRS section 12).

Report 4 in the SRS, "حركة الأتواب", is deliberately absent: it needs a roll
table and opening balances, which SRS 4.3 rules out until phase 3.
"""
import datetime
from decimal import Decimal
from io import BytesIO

from django.db.models import Avg, Count, F, Q, Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from hr.attendance import hours_for

from .models import Lay, LayLine, RemnantLog


def _base(start=None, end=None, include_backfill=False):
    qs = Lay.objects.select_related("garment_model", "bank", "team_leader")
    if not include_backfill:
        qs = qs.exclude(is_backfill=True)
    # Intersection, not a start-date match: a lay spread over two days belongs
    # to any period either day touches (SRS 5.6).
    if start:
        qs = qs.filter(end_date__gte=start)
    if end:
        qs = qs.filter(start_date__lte=end)
    return qs


def _period(start, end) -> dict:
    return {"start": start.isoformat() if start else None,
            "end": end.isoformat() if end else None}


# --- 9.1 metrage per model ------------------------------------------------

METRAGE_COLUMNS = [
    ("code", "الكود"), ("name", "الموديل"), ("fit", "القَصّة"),
    ("lays", "الفرشات"), ("theoretical_pieces", "القطع النظرية"),
    ("actual_pieces", "القطع الفعلية"), ("expected_metrage", "المتوقع"),
    ("real_metrage", "الحقيقي"), ("deviation_pct", "الانحراف %"),
]


def metrage_by_model(start=None, end=None, include_backfill=False) -> dict:
    qs = _base(start, end, include_backfill)
    rows = (
        qs.values("garment_model__code", "garment_model__name", "garment_model__fit__name")
        .annotate(
            lays=Count("id", distinct=True),
            theoretical_pieces=Sum("theoretical_pieces"),
            actual_pieces=Sum("output__actual_pieces"),
            expected_metrage=Avg("expected_metrage"),
            real_metrage=Avg("real_metrage"),
            deviation_pct=Avg("deviation_pct"),
        )
        .order_by("garment_model__code")
    )
    return {
        "title": "الميتراج لكل موديل",
        "period": _period(start, end),
        "columns": METRAGE_COLUMNS,
        "rows": [
            {
                "code": r["garment_model__code"],
                "name": r["garment_model__name"],
                "fit": r["garment_model__fit__name"] or "",
                "lays": r["lays"],
                "theoretical_pieces": r["theoretical_pieces"] or 0,
                "actual_pieces": r["actual_pieces"] or 0,
                "expected_metrage": _round(r["expected_metrage"], 4),
                "real_metrage": _round(r["real_metrage"], 4),
                "deviation_pct": _round(r["deviation_pct"], 2),
            }
            for r in rows
        ],
    }


# --- 9.2 shortage ---------------------------------------------------------

SHORTAGE_COLUMNS = [
    ("lay_id", "الفرشة"), ("date", "التاريخ"), ("code", "الكود"),
    ("team_leader", "رئيس الفريق"), ("bank", "البنك"),
    ("articles", "الخامة"), ("lots", "اللوط"),
    ("roll_length", "أطوال الأتواب"), ("consumed", "المستهلك"),
    ("remnant", "البواقي"), ("shortage", "العجز"),
]


def shortage_report(start=None, end=None, include_backfill=False) -> dict:
    """Only lays flagged over tolerance, with the articles and lots on them —
    the point is to expose a lot whose rolls are consistently short."""
    qs = _base(start, end, include_backfill).filter(has_shortage=True).order_by("-end_date")

    rows = []
    for lay in qs.prefetch_related("lines"):
        lines = list(lay.lines.all())
        rows.append({
            "lay_id": lay.pk,
            "date": lay.end_date.isoformat(),
            "code": lay.garment_model.code,
            "team_leader": lay.team_leader.full_name,
            "bank": lay.bank.name,
            "articles": " / ".join(sorted({l.article for l in lines if l.article})),
            "lots": " / ".join(sorted({l.lot_no for l in lines if l.lot_no})),
            "roll_length": lay.total_roll_length_m,
            "consumed": lay.consumed_m,
            "remnant": lay.total_remnant_m,
            "shortage": lay.fabric_shortage_m,
        })

    by_lot = {}
    for row in rows:
        for lot in [l for l in row["lots"].split(" / ") if l]:
            entry = by_lot.setdefault(lot, {"lot": lot, "lays": 0, "shortage": Decimal("0")})
            entry["lays"] += 1
            entry["shortage"] += row["shortage"]

    return {
        "title": "تقرير العجز",
        "period": _period(start, end),
        "columns": SHORTAGE_COLUMNS,
        "rows": rows,
        "total_shortage": sum((r["shortage"] for r in rows), Decimal("0")),
        "by_lot": sorted(by_lot.values(), key=lambda e: e["shortage"], reverse=True),
    }


# --- 9.3 team-leader productivity -----------------------------------------

PRODUCTIVITY_COLUMNS = [
    ("employee_code", "الكود"), ("full_name", "رئيس الفريق"),
    ("lays", "الفرشات"), ("actual_pieces", "القطع الفعلية"),
    ("hours", "الساعات"), ("pieces_per_hour", "قطعة/ساعة"),
    ("coverage", "التغطية"), ("reliable", "موثوق؟"),
]


def productivity_report(start=None, end=None, include_backfill=False) -> dict:
    """Pieces per hour from the fingerprint device.

    Days where the leader punched once and never punched out score zero hours
    and are dropped from the denominator, never averaged in as a zero — the
    man worked, he just did not punch out. Every row states the coverage it is
    built on, and says so plainly when that coverage is too thin to trust
    (SRS section 6).
    """
    qs = _base(start, end, include_backfill).filter(output__isnull=False)

    per_leader = {}
    for lay in qs.select_related("team_leader", "output"):
        entry = per_leader.setdefault(
            lay.team_leader_id,
            {"employee": lay.team_leader, "lays": 0, "pieces": 0,
             "hours": 0.0, "measured": 0, "days": 0},
        )
        entry["lays"] += 1
        entry["pieces"] += lay.output.actual_pieces
        period = hours_for(lay.team_leader.employee_code, lay.start_date, lay.end_date)
        entry["hours"] += period.total_hours
        entry["measured"] += period.measured_days
        entry["days"] += period.total_days

    rows = []
    for entry in per_leader.values():
        hours = round(entry["hours"], 2)
        coverage = round(entry["measured"] / entry["days"] * 100, 1) if entry["days"] else 0.0
        rows.append({
            "employee_code": entry["employee"].employee_code,
            "full_name": entry["employee"].full_name,
            "lays": entry["lays"],
            "actual_pieces": entry["pieces"],
            "hours": hours,
            "pieces_per_hour": round(entry["pieces"] / hours, 2) if hours > 0 else None,
            "coverage": f"{entry['measured']} من {entry['days']} يوم",
            "coverage_pct": coverage,
            "reliable": coverage >= 50 and hours > 0,
        })
    rows.sort(key=lambda r: (r["pieces_per_hour"] is None, -(r["pieces_per_hour"] or 0)))

    return {
        "title": "إنتاجية رؤساء الفرق",
        "period": _period(start, end),
        "columns": PRODUCTIVITY_COLUMNS,
        "rows": rows,
        "note": "الأيام اللي فيها بصمة واحدة بتتستبعد من الحساب، والتغطية مكتوبة لكل صف.",
    }


# --- 9.5 remnants and waste ----------------------------------------------

REMNANT_COLUMNS = [
    ("article", "الخامة"), ("lot_no", "اللوط"), ("disposition", "التصنيف"),
    ("entries", "عدد البواقي"), ("length", "الأمتار"),
]


def remnant_report(start=None, end=None, include_backfill=False) -> dict:
    qs = RemnantLog.objects.select_related("lay_line__lay")
    if not include_backfill:
        qs = qs.exclude(lay_line__lay__is_backfill=True)
    if start:
        qs = qs.filter(lay_line__lay__end_date__gte=start)
    if end:
        qs = qs.filter(lay_line__lay__start_date__lte=end)

    grouped = (
        qs.values("article", "lot_no", "disposition")
        .annotate(entries=Count("id"), length=Sum("length_m"))
        .order_by("-length")
    )
    rows = [
        {
            "article": g["article"] or "—",
            "lot_no": g["lot_no"] or "—",
            "disposition": "هالك" if g["disposition"] == "waste" else "صالح",
            "entries": g["entries"],
            "length": g["length"] or Decimal("0"),
        }
        for g in grouped
    ]
    waste = sum((r["length"] for r in rows if r["disposition"] == "هالك"), Decimal("0"))
    usable = sum((r["length"] for r in rows if r["disposition"] == "صالح"), Decimal("0"))
    return {
        "title": "البواقي والهالك",
        "period": _period(start, end),
        "columns": REMNANT_COLUMNS,
        "rows": rows,
        "total_waste_m": waste,
        "total_usable_m": usable,
    }


# --- 9.6 daily bank report ------------------------------------------------

BANK_COLUMNS = [
    ("date", "التاريخ"), ("bank", "البنك"), ("lays", "الفرشات"),
    ("plies", "إجمالي الراق"), ("theoretical_pieces", "القطع النظرية"),
    ("actual_pieces", "القطع الفعلية"), ("fabric", "الأمتار"),
]


def daily_bank_report(start=None, end=None, include_backfill=False) -> dict:
    """Production is counted on the day the lay closed, so it is not counted
    twice for a lay that ran across two days (SRS 5.6)."""
    qs = _base(start, end, include_backfill)
    grouped = (
        qs.values("end_date", "bank__name")
        .annotate(
            lays=Count("id", distinct=True),
            plies=Sum("total_plies"),
            theoretical_pieces=Sum("theoretical_pieces"),
            actual_pieces=Sum("output__actual_pieces"),
            fabric=Sum("total_roll_length_m"),
        )
        .order_by("-end_date", "bank__name")
    )
    return {
        "title": "تقرير البنوك اليومي",
        "period": _period(start, end),
        "columns": BANK_COLUMNS,
        "rows": [
            {
                "date": g["end_date"].isoformat(),
                "bank": g["bank__name"],
                "lays": g["lays"],
                "plies": g["plies"] or 0,
                "theoretical_pieces": g["theoretical_pieces"] or 0,
                "actual_pieces": g["actual_pieces"] or 0,
                "fabric": g["fabric"] or Decimal("0"),
            }
            for g in grouped
        ],
    }


# --- 9.7 entry quality ----------------------------------------------------

QUALITY_COLUMNS = [
    ("metric", "المؤشر"), ("count", "العدد"), ("pct", "النسبة %"),
]


def entry_quality_report(start=None, end=None, include_backfill=False) -> dict:
    """How disciplined the recording is — the share entered in a hurry, without
    a notebook photo, or still unnumbered (SRS 9.7)."""
    qs = _base(start, end, include_backfill)
    total = qs.count()

    def pct(n):
        return round(n / total * 100, 1) if total else 0.0

    quick = qs.filter(entry_mode=Lay.MODE_QUICK).count()
    no_photo = qs.filter(Q(sheet_image="") | Q(sheet_image__isnull=True)).count()
    awaiting = qs.filter(status=Lay.STATUS_CLOSED, output__isnull=True).count()
    mismatch = qs.filter(has_length_mismatch=True).count()
    shortage = qs.filter(has_shortage=True).count()

    metrics = [
        ("إجمالي الفرشات", total),
        ("إدخال سريع", quick),
        ("من غير صورة دفتر", no_photo),
        ("مستنية ترقيم", awaiting),
        ("فيها فرق في الأطوال", mismatch),
        ("فيها عجز", shortage),
    ]
    return {
        "title": "جودة الإدخال",
        "period": _period(start, end),
        "columns": QUALITY_COLUMNS,
        "rows": [
            {"metric": name, "count": n, "pct": 100.0 if name.startswith("إجمالي") else pct(n)}
            for name, n in metrics
        ],
        "total_lays": total,
    }


REPORTS = {
    "metrage": metrage_by_model,
    "shortage": shortage_report,
    "productivity": productivity_report,
    "remnants": remnant_report,
    "banks": daily_bank_report,
    "quality": entry_quality_report,
}


def _round(value, places):
    if value is None:
        return None
    return float(round(Decimal(str(value)), places))


# --- exports --------------------------------------------------------------

def report_xlsx(report: dict) -> BytesIO:
    """Every report shares one Excel writer, driven by its `columns` list."""
    wb = Workbook()
    ws = wb.active
    ws.title = report["title"][:31]
    ws.sheet_view.rightToLeft = True

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF", name="Arial")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    columns = report["columns"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    period = report.get("period") or {}
    span = ""
    if period.get("start") or period.get("end"):
        span = f" — من {period.get('start') or '…'} إلى {period.get('end') or '…'}"
    title = ws.cell(row=1, column=1, value=f"{report['title']}{span}")
    title.font = Font(bold=True, size=14, name="Arial")
    title.alignment = center

    for col, (_key, label) in enumerate(columns, start=1):
        c = ws.cell(row=3, column=col, value=label)
        c.fill, c.font, c.alignment, c.border = head_fill, head_font, center, border

    for i, row in enumerate(report["rows"], start=4):
        for col, (key, _label) in enumerate(columns, start=1):
            value = row.get(key)
            if isinstance(value, Decimal):
                value = float(value)
            elif isinstance(value, bool):
                value = "نعم" if value else "لا"
            c = ws.cell(row=i, column=col, value=value)
            c.alignment = center
            c.border = border

    for i, (_key, label) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(label) + 6)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def report_pdf(report: dict) -> BytesIO:
    """WeasyPrint shapes Arabic natively, so an HTML table is enough."""
    import html as html_lib

    from weasyprint import HTML

    def esc(v):
        if isinstance(v, bool):
            v = "نعم" if v else "لا"
        return html_lib.escape(str(v if v is not None else "—"))

    columns = report["columns"]
    period = report.get("period") or {}
    span = ""
    if period.get("start") or period.get("end"):
        span = f"من {esc(period.get('start') or '…')} إلى {esc(period.get('end') or '…')}"
    head = "".join(f"<th>{esc(label)}</th>" for _key, label in columns)
    body = "".join(
        "<tr>" + "".join(f"<td>{esc(row.get(key))}</td>" for key, _l in columns) + "</tr>"
        for row in report["rows"]
    )
    note = f"<p class='note'>{esc(report['note'])}</p>" if report.get("note") else ""

    html = f"""<!doctype html><html dir="rtl" lang="ar"><head><meta charset="utf-8">
<style>
  @page {{ size: A4 landscape; margin: 12mm; }}
  body {{ font-family: 'Noto Naskh Arabic','DejaVu Sans',sans-serif; font-size: 10px; }}
  h1 {{ font-size: 16px; margin: 0 0 2px; }}
  .sub {{ color: #666; font-size: 11px; margin: 0 0 10px; }}
  .note {{ color: #666; font-size: 10px; margin: 8px 0 0; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th {{ background: #1F4E78; color: #fff; padding: 5px; border: 1px solid #999; }}
  td {{ padding: 4px 5px; border: 1px solid #ccc; text-align: center; }}
  tr:nth-child(even) td {{ background: #f6f8fa; }}
</style></head><body>
<h1>{esc(report['title'])}</h1><p class="sub">{span}</p>
<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>
{note}
</body></html>"""

    buf = BytesIO()
    HTML(string=html).write_pdf(buf)
    buf.seek(0)
    return buf
